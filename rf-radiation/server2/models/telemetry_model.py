"""
Telemetry Model - Handles telemetry data loading and management
"""
from typing import List, Dict
from datetime import datetime

class TelemetryModel:
    """Model for telemetry data"""
    
    def __init__(self):
        self.data: List[Dict] = []
        self.current_file: str = None

    def load_from_xlsx(self, filepath: str) -> bool:
        """Load telemetry from XLSX file"""
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(filepath, read_only=True)
            ws = wb.active
            
            self.data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[3] and row[4] and row[5]:
                    roll_deg = float(row[10]) if len(row) > 10 and row[10] else 0.0
                    pitch_deg = float(row[11]) if len(row) > 11 and row[11] else 0.0
                    yaw_deg = float(row[12]) if len(row) > 12 and row[12] else 0.0
                    
                    self.data.append({
                        "timestamp": row[0].isoformat() if hasattr(row[0], 'isoformat') else str(row[0]),
                        "lat": float(row[3]),
                        "lon": float(row[4]),
                        "alt": float(row[5]),
                        "ground_speed": float(row[6] or 0),
                        "vertical_speed": float(row[7] or 0),
                        "num_sats": int(row[8] or 0),
                        "pressure": float(row[9] or 0),
                        "roll_deg": roll_deg,
                        "pitch_deg": pitch_deg,
                        "yaw_deg": yaw_deg,
                    })
            return len(self.data) > 0
        except Exception as e:
            print(f"[TELEMETRY] Error loading file: {e}")
            return False

    def get_point(self, index: int) -> Dict:
        """Get telemetry point at index"""
        if 0 <= index < len(self.data):
            return self.data[index]
        return None

    def get_all(self) -> List[Dict]:
        """Get all telemetry data"""
        return self.data

    def get_count(self) -> int:
        """Get number of telemetry points"""
        return len(self.data)

    def clear(self):
        """Clear telemetry data"""
        self.data = []
