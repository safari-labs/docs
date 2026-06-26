#!/usr/bin/env python3
from openpyxl import load_workbook

wb = load_workbook('/home/ouazo/Desktop/ASSFID/server2/telemetry_uploads/tm2.xlsx', read_only=True)
ws = wb.active

print(f"Sheet name: {ws.title}")
print(f"\nTotal rows: {ws.max_row}")
print("\nFirst 3 rows:")

for i, row in enumerate(ws.iter_rows(max_row=3, values_only=True)):
    print(f"Row {i}: {row[:10]}")
